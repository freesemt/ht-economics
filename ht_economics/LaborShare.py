"""
LaborShare.py
"""
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

def load_laborshare_data(bookfile):
    wb = load_workbook(bookfile)
    ws = wb.active
    for k, row in enumerate(ws.iter_rows(max_col=50)):
        values = [cell.value for cell in row]
        # print([k], values)
        first_value = values[0]
        if first_value is not None and first_value.find("雇用者報酬") >= 0:
            size = last_values[1:].index(None)
            years = last_values[1:1+size]
            wages = values[1:1+size]
        elif first_value is not None and first_value.find("国民所得（要素費用表示）") >= 0:
            incomes1 = values[1:1+size]
        elif first_value is not None and first_value.find("国民所得（市場価格表示）") >= 0:
            incomes2 = values[1:1+size]
            break
        last_values = values
    wb.close()
    return years, wages, incomes1, incomes2

def plot_laborshare(years, wages, incomes1, incomes2):
    fig, (ax1, ax2) = plt.subplots(ncols=2, figsize=(12, 5))
    ax1.set_xlabel("年度")
    ax1.set_ylabel("報酬 or 所得")
    ax1.set_title("雇用者報酬と国民所得")
    ax1.plot(years, wages, label="雇用者報酬")
    ax1.plot(years, incomes1, label="国民所得（要素費用表示）")
    ax1.plot(years, incomes2, label="国民所得（市場価格表示）")
    ax1.set_ylim(0, 500000)
    ax1.legend()

    ax2.set_title("労働分配率")
    ax2.set_xlabel("年度")
    ax2.set_ylabel("労働分配率")
    wages = np.array(wages, dtype=np.float64)
    incomes1 = np.array(incomes1, dtype=np.float64)
    incomes2 = np.array(incomes2, dtype=np.float64)
    ax2.plot(years, wages / incomes1 * 100, label="雇用者報酬 / 国民所得（要素費用表示）")
    ax2.plot(years, wages / incomes2 * 100, label="雇用者報酬 / 国民所得（市場価格表示）")
    ax2.legend()
    ax2.set_ylim(0, 100)
    fig.tight_layout()